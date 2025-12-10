import matplotlib.pyplot as plt
from scipy.signal import savgol_filter, cheby2, filtfilt
from scipy.interpolate import interp1d
from scipy.optimize import curve_fit
import numpy as np
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.preprocessing import StandardScaler
import openpyxl
from tools_plus import *
import os


# Gaussian function definitions
def gaussian1(x, a, b, c):
    """Single Gaussian function"""
    return a * np.exp(-(x - b) ** 2 / (2 * c ** 2))


def gaussian3(x, a1, a2, a3, b1, b2, b3, c1, c2, c3):
    """Three Gaussian superposition function"""
    return (gaussian1(x, a1, b1, c1) +
            gaussian1(x, a2, b2, c2) +
            gaussian1(x, a3, b3, c3))


def calc_residual(data1, data2):
    sum = 0
    for i in range(0, len(data1)):
        sum = sum + np.power((data1[i] - data2[i]), 2)
    result = sum / len(data1)
    return result


# Create or load result Excel file
result_file = '2.Gaussian_optimal_parameters.xlsx'
if os.path.exists(result_file):
    book2 = openpyxl.load_workbook(result_file)
else:
    book2 = openpyxl.Workbook()

if 'Sheet1' in book2.sheetnames:
    sheet2 = book2['Sheet1']
else:
    sheet2 = book2.active
    sheet2.title = 'Sheet1'
    # Write headers
    headers = ['Name', 'a1', 'a2', 'a3', 'b1', 'b2', 'b3', 'c1', 'c2', 'c3', 'Residual']
    for col, header in enumerate(headers, 1):
        sheet2.cell(row=1, column=col, value=header)



bounds_df = pd.read_excel('1_Data_interval_boundaries.xlsx')
bounds_dict = {row['Feature']: {'lower': row['Lower Bound'], 'upper': row['Upper Bound']}
               for _, row in bounds_df.iterrows()}

# Load original PPG feature Excel file
book1 = openpyxl.load_workbook('8D_KF.xlsx')
sheet1 = book1['Sheet1']
sheet1_max_row = sheet1.max_row

for row in range(2, sheet1_max_row + 1):
    name = sheet1.cell(row, 1).value  # Subject ID
    start = sheet1.cell(row, 2).value  # Signal start point
    end = sheet1.cell(row, 3).value  # Signal end point
    file_name = name + '.txt'
    path = '../../../dataset/' + file_name  # PPG data file path

    ########################################## Preprocessing pipeline #######################################
    # 1. Read raw PPG signal
    ppg = Read_raw_ppg_signal(path)

    # 2. Apply Savitzky-Golay filter for smoothing
    filter_ppg = savgol_filter(ppg, 201, 2)

    # 3. Extract single heartbeat cycle
    ppg_0 = filter_ppg[start:end]

    # 4. Remove baseline drift
    detrend_filter_ppg_0 = detrend(ppg_0)
    y = detrend_filter_ppg_0

    x = np.linspace(1, len(ppg_0), len(ppg_0))  # Create x-axis coordinates

    # Modified parameter settings section
    p_a1 = 0.8 * bounds_dict['h1']['lower']
    low_a1 = 0.8 * bounds_dict['h1']['lower']
    upper_a1 = bounds_dict['h1']['upper']

    # Use values read from Excel to set parameters
    p_a2 = 0.8 * bounds_dict['h2']['lower']
    low_a2 = 0.8 * bounds_dict['h2']['lower']
    upper_a2 = bounds_dict['h2']['upper']

    p_a3 = 0.8 * bounds_dict['h3']['lower']
    low_a3 = 0.8 * bounds_dict['h3']['lower']
    upper_a3 = bounds_dict['h3']['upper']

    p_b1 = bounds_dict['t1']['lower']
    low_b1 = bounds_dict['t1']['lower'] - 0.1 * len(y)
    upper_b1 = bounds_dict['t1']['upper'] + 0.1 * len(y)

    p_b2 = bounds_dict['t2']['lower']
    low_b2 = bounds_dict['t2']['lower'] - 0.1 * len(y)
    upper_b2 = bounds_dict['t2']['upper'] + 0.1 * len(y)

    p_b3 = bounds_dict['t3']['lower']
    low_b3 = bounds_dict['t3']['lower'] - 0.1 * len(y)
    upper_b3 = bounds_dict['t3']['upper'] + 0.1 * len(y)

    p_c1 = bounds_dict['Tsys']['lower'] / 6
    low_c1 = 0.5 * bounds_dict['Tsys']['lower'] / 6
    upper_c1 = 1.5 * bounds_dict['Tsys']['upper'] / 6

    p_c2 = bounds_dict['Tsys']['lower'] / 6
    low_c2 = 0.5 * bounds_dict['Tsys']['lower'] / 6
    upper_c2 = 1.5 * bounds_dict['Tsys']['upper'] / 6

    p_c3 = bounds_dict['Tdia']['lower'] / 6
    low_c3 = 0.5 * bounds_dict['Tdia']['lower'] / 6
    upper_c3 = 1.5 * bounds_dict['Tdia']['upper'] / 6

    p = [p_a1, p_a2, p_a3, p_b1, p_b2, p_b3, p_c1, p_c2, p_c3]
    low = [low_a1, low_a2, low_a3, low_b1, low_b2, low_b3, low_c1, low_c2, low_c3]
    upper = [upper_a1, upper_a2, upper_a3, upper_b1, upper_b2, upper_b3, upper_c1, upper_c2, upper_c3]

    popt, pcov = curve_fit(gaussian3, x, y, p0=p, bounds=(low, upper))
    fit_signal = gaussian3(x, *popt)

    residual = calc_residual(y, fit_signal)
    print(residual)

    # Calculate three independent Gaussian components
    g1 = gaussian1(x, popt[0], popt[3], popt[6])
    g2 = gaussian1(x, popt[1], popt[4], popt[7])
    g3 = gaussian1(x, popt[2], popt[5], popt[8])

    # Plot fitting results
    plt.figure(figsize=(16, 10), dpi=300)
    plt.plot(x, y, label='Normalized PPG', linewidth=3)
    plt.plot(x, gaussian3(x, *popt), label='Fitted PPG', linestyle='--', linewidth=5)
    plt.plot(x, g1, label='Gaussian Component 1', linewidth=3)
    plt.plot(x, g2, label='Gaussian Component 2', linewidth=3)
    plt.plot(x, g3, label='Gaussian Component 3', linewidth=3)
    plt.title(f'Gaussian Fitting Results - {name}')
    plt.legend(fontsize=16, frameon=False)

    # Create output directory (if not exists)
    os.makedirs('../gaussian_fitting_plots', exist_ok=True)
    plt.savefig(f'../gaussian_fitting_plots/{name}.png')
    plt.close()

    # Save results to Excel
    result_row = [name] + list(popt) + [residual]
    sheet2.append(result_row)

    print(f"Successfully processed subject: {name}")

# Save result Excel file
book2.save(result_file)
book1.close()
book2.close()