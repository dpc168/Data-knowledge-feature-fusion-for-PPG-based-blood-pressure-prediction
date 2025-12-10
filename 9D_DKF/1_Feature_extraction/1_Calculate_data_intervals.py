import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Read Excel file
file_path = '8D_KF.xlsx'
df = pd.read_excel(file_path)

# Specify feature columns to calculate
features = ['h1','h2', 'h3', 't1', 't2', 't3', 'Tsys', 'Tdia']

# Calculate 99% data interval for each feature
results = []

for i, feature in enumerate(features, 1):
    # Extract feature column data
    data = df[feature].dropna()  # Remove NaN values

    # 99% data interval (0.5%-99.5%)
    lower_bound = data.quantile(0.005)
    upper_bound = data.quantile(0.995)

    # Save results
    results.append({
        'Feature': feature,
        'Lower Bound': lower_bound,
        'Upper Bound': upper_bound
    })

# Write statistical results to Excel
output_path = '1_Data_interval_boundaries.xlsx'
result_df = pd.DataFrame(results)
result_df.to_excel(output_path, index=False)

# Adjust Excel formatting
wb = load_workbook(output_path)
ws = wb.active

# Set styles
alignment = Alignment(horizontal='center', vertical='center')

for col in ws.columns:
    column_letter = get_column_letter(col[0].column)
    ws.column_dimensions[column_letter].width = 20
    for cell in col:
        cell.alignment = alignment

wb.save(output_path)

print(f"Statistical results saved to {output_path}")
print("\n data interval results for each feature:")
for result in results:
    print(f"{result['Feature']}: [{result['Lower Bound']:.4f}, {result['Upper Bound']:.4f}]")