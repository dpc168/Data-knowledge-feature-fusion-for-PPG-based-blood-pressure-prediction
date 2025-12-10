import matplotlib.pyplot as plt
from tools_plus import *
from scipy.signal import savgol_filter
import openpyxl
import pandas as pd
import os
import numpy as np

# Create or load result file
result_file = ' 01_result.xlsx'

# Initialize result DataFrame
if os.path.exists(result_file):
    try:
        result_df = pd.read_excel(result_file)
        # Ensure DataFrame has correct columns
        required_columns = ['Name', 't1', 'h1', 't2', 'h2', 'tsys', 't3', 'h3', 'T', 'Tdia']
        if not all(col in result_df.columns for col in required_columns):
            result_df = pd.DataFrame(columns=required_columns)
    except:
        result_df = pd.DataFrame(columns=['Name', 't1', 'h1', 't2', 'h2', 'tsys', 't3', 'h3', 'T', 'Tdia'])
else:
    result_df = pd.DataFrame(columns=['Name', 't1', 'h1', 't2', 'h2', 'tsys', 't3', 'h3', 'T', 'Tdia'])

book1 = openpyxl.load_workbook('raw_PPG.xlsx')
sheet1 = book1['Sheet1']
sheet1_max_row = sheet1.max_row

dataset_root = '../../dataset\\'


class PointSelector:
    def __init__(self, ax_ppg4, time, ppg_4, ppg4_marked_points, name, detrend_ppg_signal):
        self.ax_ppg4 = ax_ppg4
        self.time = time
        self.ppg_4 = ppg_4
        self.ppg4_marked_points = ppg4_marked_points
        self.name = name
        self.selected_points = []  # Storage format: (ppg4_x, detrend_y)
        self.title = ax_ppg4.get_title()
        self.detrend_ppg_signal = detrend_ppg_signal  # detrend_ppg signal for getting height
        self.ppg_len = len(detrend_ppg_signal)  # PPG signal length

        # Connect click event
        self.cid = ax_ppg4.figure.canvas.mpl_connect('button_press_event', self.on_click)

        # Add confirmation button
        self.done_button = ax_ppg4.figure.text(0.85, 0.02, 'Done',
                                               bbox=dict(boxstyle="round,pad=0.3", facecolor='lightblue'),
                                               picker=True)

    def on_click(self, event):
        # Check if done button was clicked
        if event.inaxes is None and self.done_button.contains(event)[0]:
            if len(self.selected_points) == 4:
                self.save_points()
                plt.close(event.canvas.figure)
            else:
                print(f"Please select 4 points, currently selected {len(self.selected_points)} points")
            return

        if event.inaxes != self.ax_ppg4:
            return

        # Find the closest point in marked points
        min_dist = float('inf')
        closest_point = None

        for point in self.ppg4_marked_points:
            x, y = point
            dist = ((x - event.xdata) ** 2 + (y - event.ydata) ** 2) ** 0.5
            if dist < min_dist and dist < 10:  # 10 pixel tolerance
                min_dist = dist
                closest_point = point

        if closest_point and closest_point not in [p[0] for p in self.selected_points]:
            # Get PPG4 x coordinate
            ppg4_x = closest_point[0]

            # Find corresponding detrend_ppg y value through PPG4 x coordinate
            index = int(ppg4_x)
            if index < len(self.detrend_ppg_signal):
                detrend_y = self.detrend_ppg_signal[index]

                # Save format: (ppg4_x, detrend_y)
                self.selected_points.append((ppg4_x, detrend_y))

                # Mark selected point on the graph
                # Mark clicked position on PPG4
                self.ax_ppg4.plot(ppg4_x, closest_point[1], 'g*', markersize=12, markeredgewidth=2,
                                  markeredgecolor='red', label=f'Selected point {len(self.selected_points)}')

                # Update title
                self.ax_ppg4.set_title(f"Please select 4 feature points on PPG4 for {self.name}\n(Selected: {len(self.selected_points)}/4)",
                                       color='red')

                # Immediately save current point
                self.save_current_point(len(self.selected_points))

                # Redraw
                event.canvas.draw()

                print(f"Selected point {len(self.selected_points)}: PPG4_x={ppg4_x:.2f}, detrend_y={detrend_y:.4f}")

                # Automatically close if 4 points have been selected
                if len(self.selected_points) == 4:
                    print("4 points selected, closing automatically...")
                    plt.close(event.canvas.figure)

    def save_current_point(self, point_num):
        """Save currently selected single point"""
        global result_df

        if point_num <= len(self.selected_points):
            ppg4_x, detrend_y = self.selected_points[point_num - 1]

            # Prepare data
            row_data = {
                'Name': self.name,
                'T': self.ppg_len  # T is PPG signal length
            }

            # Assign fields based on point order
            if point_num == 1:
                row_data['t1'] = ppg4_x
                row_data['h1'] = detrend_y
            elif point_num == 2:
                row_data['t2'] = ppg4_x
                row_data['h2'] = detrend_y
            elif point_num == 3:
                row_data['tsys'] = ppg4_x  # Third point is tsys
                # Calculate Tdia = T - tsys
                row_data['Tdia'] = self.ppg_len - ppg4_x
            elif point_num == 4:
                row_data['t3'] = ppg4_x
                row_data['h3'] = detrend_y

            # Check if record for this Name already exists
            if not result_df.empty and 'Name' in result_df.columns and self.name in result_df[
                'Name'].values:
                # Update existing record
                for col, value in row_data.items():
                    if col != 'Name':  # Don't update Name
                        result_df.loc[result_df['Name'] == self.name, col] = value
            else:
                # Create new record if this is the first point
                if point_num == 1:
                    # Initialize all fields as None
                    new_row = {
                        'Name': self.name,
                        't1': ppg4_x,
                        'h1': detrend_y,
                        't2': None,
                        'h2': None,
                        'tsys': None,
                        't3': None,
                        'h3': None,
                        'T': self.ppg_len,
                        'Tdia': None
                    }
                    result_df = pd.concat([result_df, pd.DataFrame([new_row])], ignore_index=True)
                else:
                    # Find record for this Name and update
                    if self.name in result_df['Name'].values:
                        for col, value in row_data.items():
                            if col != 'Name':
                                result_df.loc[result_df['Name'] == self.name, col] = value

            # Save to Excel
            try:
                result_df.to_excel(result_file, index=False)
                print(f"Saved point {point_num}: PPG4_x={ppg4_x:.2f}, detrend_y={detrend_y:.4f}")
            except Exception as e:
                print(f"Error saving file: {e}")

    def save_points(self):
        """Save all points (for done button)"""
        global result_df

        # Ensure all 4 points have been saved
        if len(self.selected_points) == 4:
            t1, h1 = self.selected_points[0]
            t2, h2 = self.selected_points[1]
            tsys = self.selected_points[2][0]  # Third point only takes x coordinate
            t3, h3 = self.selected_points[3]

            # Calculate Tdia
            Tdia = self.ppg_len - tsys

            # Update final record
            if self.name in result_df['Name'].values:
                result_df.loc[result_df['Name'] == self.name, 't1'] = t1
                result_df.loc[result_df['Name'] == self.name, 'h1'] = h1
                result_df.loc[result_df['Name'] == self.name, 't2'] = t2
                result_df.loc[result_df['Name'] == self.name, 'h2'] = h2
                result_df.loc[result_df['Name'] == self.name, 'tsys'] = tsys
                result_df.loc[result_df['Name'] == self.name, 't3'] = t3
                result_df.loc[result_df['Name'] == self.name, 'h3'] = h3
                result_df.loc[result_df['Name'] == self.name, 'T'] = self.ppg_len
                result_df.loc[result_df['Name'] == self.name, 'Tdia'] = Tdia

                # Save to Excel
                result_df.to_excel(result_file, index=False)
                print(f"Completed saving all 4 points for {self.name}")
                print(
                    f"t1={t1:.2f}, h1={h1:.4f}, t2={t2:.2f}, h2={h2:.4f}, tsys={tsys:.2f}, t3={t3:.2f}, h3={h3:.4f}, T={self.ppg_len}, Tdia={Tdia:.2f}")
            else:
                print(f"Error: Record for {self.name} not found")


for row in range(2, sheet1_max_row + 1):
    subject_id = sheet1.cell(row, 1).value
    name = str(subject_id)
    start = sheet1.cell(row, 2).value
    end = sheet1.cell(row, 3).value

    # 1. Read raw signal
    ppg = Read_raw_ppg_signal(dataset_root + subject_id + '.txt')
    ppg_signal = ppg[start:end]  # Store original PPG signal
    # 2. Filter processing
    filter_ppg = savgol_filter(ppg, 101, 2)

    # 3. Extract single cycle
    ppg = filter_ppg[start:end]

    # 4. Detrend
    detrend_ppg = detrend(ppg)
    detrend_ppg_signal = detrend_ppg  # Save detrend signal for height calculation

    # 5. Normalization
    ppg = min_max_normalization(ppg)
    ppg_signal = min_max_normalization(ppg_signal)
    # detrend_ppg = min_max_normalization(detrend_ppg)

    # Calculate PPG fourth derivative
    ppg_0 = ppg  # Save PPG0 signal
    ppg_1 = np.gradient(ppg_0)
    ppg_1 = savgol_filter(ppg_1, 101, 2)  # First derivative
    ppg_2 = np.gradient(ppg_1)
    ppg_2 = savgol_filter(ppg_2, 121, 2)  # Second derivative
    ppg_3 = np.gradient(ppg_2)
    ppg_3 = savgol_filter(ppg_3, 91, 2)  # Third derivative
    ppg_4 = np.gradient(ppg_3)
    ppg_4 = savgol_filter(ppg_4, 91, 2)  # Fourth derivative

    time = np.linspace(0, len(ppg), len(ppg))

    # Detect all peaks and valleys of PPG4
    ppg4_max_array, ppg4_max_index_array = local_max(ppg_4, 51)
    ppg4_min_array, ppg4_min_index_array = local_min(ppg_4, 51)

    # Store coordinates of all PPG4 marked points
    ppg4_marked_points = []
    for i in range(len(ppg4_max_index_array)):
        ppg4_marked_points.append((ppg4_max_index_array[i], ppg4_max_array[i]))
    for i in range(len(ppg4_min_index_array)):
        ppg4_marked_points.append((ppg4_min_index_array[i], ppg4_min_array[i]))

    ########################################## Plot graphs #####################################
    fig, ax = plt.subplots(1, 2, figsize=(15, 6), dpi=100)

    # First graph: detrend_ppg
    ax[0].plot(time, detrend_ppg, color='black', label='detrend_PPG')
    ax[0].legend(loc='upper right')
    ax[0].set_title(f"detrend_PPG signal for {name}", fontsize=12)
    ax[0].set_xlabel('Time')
    ax[0].set_ylabel('Amplitude')

    # Second graph: PPG4 with detrend_ppg (operation area)
    # Plot detrend_ppg
    ax[1].plot(time, detrend_ppg, color='black', label='detrend_PPG', alpha=0.7)
    ax[1].set_xlabel('Time')
    ax[1].set_ylabel('detrend_PPG Amplitude', color='black')
    ax[1].tick_params(axis='y', labelcolor='black')

    # Create second y-axis for PPG4
    ax_ppg4 = ax[1].twinx()
    ax_ppg4.plot(time, ppg_4, color='orange', label='PPG4', alpha=0.8)

    # Mark all peaks and valleys of PPG4
    ax_ppg4.plot(ppg4_max_index_array, ppg4_max_array, 'ro', markersize=8, label='PPG4 Peaks')
    ax_ppg4.plot(ppg4_min_index_array, ppg4_min_array, 'bo', markersize=8, label='PPG4 Valleys')

    # Set PPG4 y-axis label
    ax_ppg4.set_ylabel('PPG4 Amplitude', color='orange')
    ax_ppg4.tick_params(axis='y', labelcolor='orange')

    # Combine legends
    lines_detrend = ax[1].get_lines()
    lines_ppg4 = ax_ppg4.get_lines()
    lines = lines_detrend + lines_ppg4
    labels = [line.get_label() for line in lines]
    ax[1].legend(lines, labels, loc='upper right')

    ax[1].set_title(f"Please select 4 feature points on PPG4 for {name}\n(Selected: 0/4)", color='red', fontsize=12)

    # Set overall operation instructions
    plt.figtext(0.5, 0.01,
                "Instructions: Select 4 PPG4 feature points in the right graph, each point is automatically saved (saving PPG4 x-coordinate and detrend_ppg y-coordinate)",
                ha='center', fontsize=10, color='red')

    plt.tight_layout()

    # Create point selector - pass detrend_ppg_signal for getting height
    selector = PointSelector(ax_ppg4, time, ppg_4, ppg4_marked_points, name, detrend_ppg_signal)

    plt.show()

print("All processing completed!")