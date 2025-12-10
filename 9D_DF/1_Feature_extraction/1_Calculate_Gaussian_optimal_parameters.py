import matplotlib.pyplot as plt
from scipy.signal import savgol_filter
import numpy as np
import openpyxl
from tools_plus import *
import os
from scipy.optimize import curve_fit


def gaussian(x, a, b, c):
    """Gaussian function for fitting"""
    return a * np.exp(-((x - b) ** 2) / (2 * c ** 2))


def multi_gaussian(x, a1, b1, c1, a2, b2, c2, a3, b3, c3):
    """Sum of three Gaussian functions"""
    return (gaussian(x, a1, b1, c1) +
            gaussian(x, a2, b2, c2) +
            gaussian(x, a3, b3, c3))

# Load original Excel file
book1 = openpyxl.load_workbook('raw_PPG.xlsx')
sheet1 = book1['Sheet1']
sheet1_max_row = sheet1.max_row

# Create new Excel file for saving Gaussian features
output_book = openpyxl.Workbook()
output_sheet = output_book.active
output_sheet.append(['Name', 'a1', 'b1', 'c1', 'a2', 'b2', 'c2', 'a3', 'b3', 'c3'])

# Create directory for saving images
output_dir = 'Gaussian_Fitting_Plots'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

for row in range(2, sheet1_max_row + 1):
    name = sheet1.cell(row, 1).value  # Get name
    start = sheet1.cell(row, 2).value  # Get start point
    end = sheet1.cell(row, 3).value  # Get end point
    file_name = name + '.txt'  # Build file name
    path = '../../dataset/' + file_name  # Build file path

    ########################################## Preprocessing #######################################
    # 1. Read raw PPG signal data
    ppg = Read_raw_ppg_signal(path)

    # 2. Apply Savitzky-Golay filter for signal smoothing
    filter_ppg = savgol_filter(ppg, 201, 2)

    # 3. Extract single cycle PPG signal
    ppg = filter_ppg[start:end]

    # 4. Remove baseline drift
    ppg = detrend(ppg)


    ########################################## End Preprocessing #######################################

    # Create time axis
    x = np.arange(len(ppg))

    initial_guess = [
        1.0, len(ppg) / 4, len(ppg) / 10,
        0.5, 2*len(ppg) / 5, len(ppg) / 10,
        0.2, 6*len(ppg)/ 7, len(ppg) / 10
    ]

    try:

        # Use curve fitting to fit three Gaussian functions model
        popt, pcov = curve_fit(multi_gaussian, x, ppg, p0 = initial_guess, maxfev=80000)

        print(f"{name} three Gaussian fitting successful")

        # Save fitted parameters to output sheet
        output_sheet.append([name] + list(popt))

        # Create figure
        plt.figure(figsize=(12, 8))

        # Plot original PPG signal
        plt.plot(x, ppg, 'b-', label='Original PPG Signal', linewidth=2)

        # Plot combined fitted curve
        plt.plot(x, multi_gaussian(x, *popt), 'k--', label='Combined Fitted Curve', linewidth=2)

        # Plot three individual Gaussian components
        plt.plot(x, gaussian(x, *popt[0:3]), 'r:', label='Gaussian Component 1', linewidth=1.5)
        plt.plot(x, gaussian(x, *popt[3:6]), 'g:', label='Gaussian Component 2', linewidth=1.5)
        plt.plot(x, gaussian(x, *popt[6:9]), 'm:', label='Gaussian Component 3', linewidth=1.5)

        # Add legend and title
        plt.legend(fontsize=10)
        plt.title(f'Gaussian Fitting Results for {name} (Three Components)', fontsize=12)
        plt.xlabel('Time (Sample Points)', fontsize=10)
        plt.ylabel('Amplitude', fontsize=10)
        plt.grid(True, linestyle='--', alpha=0.6)

        # Adjust layout
        plt.tight_layout()

        # Save figure to file
        plot_filename = os.path.join(output_dir, f'{name}_gaussian_fit.png')
        plt.savefig(plot_filename, dpi=300, bbox_inches='tight')
        plt.close()  # Close figure to free memory

    except Exception as e:
        print(f"{name} three Gaussian fitting failed: {e}")
        # Record NaN values directly
        output_sheet.append([name] + [np.nan] * 9)

# Save output file
output_book.save('1.Gaussian_optimal_parameters.xlsx')
print("All samples processed!")