import matplotlib.pyplot as plt
from scipy.signal import savgol_filter, cheby2, filtfilt
from scipy.interpolate import interp1d
from scipy.optimize import curve_fit
import numpy as np
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.preprocessing import StandardScaler
import openpyxl
from tools_plus import *
import os


# Gaussian function definitions
def gaussian1(x, a, b, c):
    """Single Gaussian function"""
    return a * np.exp(-(x - b) ** 2 / (2 * c ** 2))


def gaussian3(x, a1, a2, a3, b1, b2, b3, c1, c2, c3):
    """Three Gaussian superposition function"""
    return (gaussian1(x, a1, b1, c1) +
            gaussian1(x, a2, b2, c2) +
            gaussian1(x, a3, b3, c3))


# Load Excel file containing Gaussian initial parameters
gauss_params_book = openpyxl.load_workbook('1_Guassian_boundary.xlsx')
gauss_params_sheet = gauss_params_book['Guassian_boundary']

# Create dictionary to store Gaussian parameters for each subject
gauss_params = {}
for row in range(2, gauss_params_sheet.max_row + 1):
    name = gauss_params_sheet.cell(row, 1).value
    params = {
        'start_a1': gauss_params_sheet.cell(row, 2).value,  # Initial parameter a1
        'start_a2': gauss_params_sheet.cell(row, 3).value,  # Initial parameter a2
        'start_a3': gauss_params_sheet.cell(row, 4).value,  # Initial parameter a3
        'start_b1': gauss_params_sheet.cell(row, 5).value,  # Initial parameter b1
        'start_b2': gauss_params_sheet.cell(row, 6).value,  # Initial parameter b2
        'start_b3': gauss_params_sheet.cell(row, 7).value,  # Initial parameter b3
        'start_c1': gauss_params_sheet.cell(row, 8).value,  # Initial parameter c1
        'start_c2': gauss_params_sheet.cell(row, 9).value,  # Initial parameter c2
        'start_c3': gauss_params_sheet.cell(row, 10).value,  # Initial parameter c3
        'lower_a1': gauss_params_sheet.cell(row, 11).value,  # Parameter a1 lower bound
        'lower_a2': gauss_params_sheet.cell(row, 12).value,  # Parameter a2 lower bound
        'lower_a3': gauss_params_sheet.cell(row, 13).value,  # Parameter a3 lower bound
        'lower_b1': gauss_params_sheet.cell(row, 14).value,  # Parameter b1 lower bound
        'lower_b2': gauss_params_sheet.cell(row, 15).value,  # Parameter b2 lower bound
        'lower_b3': gauss_params_sheet.cell(row, 16).value,  # Parameter b3 lower bound
        'lower_c1': gauss_params_sheet.cell(row, 17).value,  # Parameter c1 lower bound
        'lower_c2': gauss_params_sheet.cell(row, 18).value,  # Parameter c2 lower bound
        'lower_c3': gauss_params_sheet.cell(row, 19).value,  # Parameter c3 lower bound
        'upper_a1': gauss_params_sheet.cell(row, 20).value,  # Parameter a1 upper bound
        'upper_a2': gauss_params_sheet.cell(row, 21).value,  # Parameter a2 upper bound
        'upper_a3': gauss_params_sheet.cell(row, 22).value,  # Parameter a3 upper bound
        'upper_b1': gauss_params_sheet.cell(row, 23).value,  # Parameter b1 upper bound
        'upper_b2': gauss_params_sheet.cell(row, 24).value,  # Parameter b2 upper bound
        'upper_b3': gauss_params_sheet.cell(row, 25).value,  # Parameter b3 upper bound
        'upper_c1': gauss_params_sheet.cell(row, 26).value,  # Parameter c1 upper bound
        'upper_c2': gauss_params_sheet.cell(row, 27).value,  # Parameter c2 upper bound
        'upper_c3': gauss_params_sheet.cell(row, 28).value  # Parameter c3 upper bound
    }
    gauss_params[name] = params

# Load original PPG feature Excel file
book1 = openpyxl.load_workbook('8D_KF.xlsx')
sheet1 = book1['Sheet1']
sheet1_max_row = sheet1.max_row

# Create new workbook to store Gaussian features
output_book = openpyxl.Workbook()
output_sheet = output_book.active
output_sheet.title = "Sheet1"

# Write headers
headers = ["Name", "a1", "a2", "a3", "b1", "b2", "b3", "c1", "c2", "c3"]
for col, header in enumerate(headers, 1):
    output_sheet.cell(1, col).value = header

for row in range(2, sheet1_max_row + 1):
    name = sheet1.cell(row, 1).value  # Subject ID
    start = sheet1.cell(row, 2).value  # Signal start point
    end = sheet1.cell(row, 3).value  # Signal end point
    file_name = name + '.txt'
    path = '../../dataset/' + file_name  # PPG data file path

    ########################################## Preprocessing pipeline #######################################
    # 1. Read raw PPG signal
    ppg = Read_raw_ppg_signal(path)

    # 2. Smooth using Savitzky-Golay filter
    filter_ppg = savgol_filter(ppg, 201, 2)

    # 3. Extract single heartbeat cycle
    ppg_0 = filter_ppg[start:end]

    # 4. Remove baseline drift
    detrend_filter_ppg_0 = detrend(ppg_0)
    y = detrend_filter_ppg_0

    # 5. Normalization
    # y = min_max_normalization(detrend_filter_ppg_0)
    x = np.linspace(1, len(ppg_0), len(ppg_0))  # Create x-axis coordinates

    # Get initial parameters and boundary conditions for this subject
    if name in gauss_params:
        params = gauss_params[name]
        # Initial parameter array
        p0 = [
            params['start_a1'], params['start_a2'], params['start_a3'],
            params['start_b1'], params['start_b2'], params['start_b3'],
            params['start_c1'], params['start_c2'], params['start_c3']
        ]
        # Parameter lower bounds
        lower_bounds = [
            params['lower_a1'], params['lower_a2'], params['lower_a3'],
            params['lower_b1'], params['lower_b2'], params['lower_b3'],
            params['lower_c1'], params['lower_c2'], params['lower_c3']
        ]
        # Parameter upper bounds
        upper_bounds = [
            params['upper_a1'], params['upper_a2'], params['upper_a3'],
            params['upper_b1'], params['upper_b2'], params['upper_b3'],
            params['upper_c1'], params['upper_c2'], params['upper_c3']
        ]

        # Perform Gaussian fitting with boundary constraints
        try:
            popt, pcov = curve_fit(gaussian3, x, y, p0=p0, bounds=(lower_bounds, upper_bounds))
            print(popt)

            # Write fitting results to output sheet
            output_sheet.cell(row, 1).value = name
            for i in range(9):
                output_sheet.cell(row, i + 2).value = popt[i]

            # Calculate three independent Gaussian components
            g1 = gaussian1(x, popt[0], popt[3], popt[6])
            g2 = gaussian1(x, popt[1], popt[4], popt[7])
            g3 = gaussian1(x, popt[2], popt[5], popt[8])

            # Plot fitting results
            plt.figure(figsize=(16, 10), dpi=300)
            plt.plot(x, y, label='Normalized PPG', linewidth=3)
            plt.plot(x, gaussian3(x, *popt), label='Fitted PPG', linestyle='--', linewidth=5)
            plt.plot(x, g1, label='Gaussian Component 1', linewidth=3)
            plt.plot(x, g2, label='Gaussian Component 2', linewidth=3)
            plt.plot(x, g3, label='Gaussian Component 3', linewidth=3)
            plt.legend(fontsize=16, frameon=False)

            # Create output directory (if it doesn't exist)
            os.makedirs('../gaussian_fitting_plots', exist_ok=True)
            plt.savefig(f'../gaussian_fitting_plots/{name}.png')
            plt.close()

            print(f"Successfully processed subject: {name}")

        except Exception as e:
            print(f"Gaussian fitting failed - {name}: {str(e)}")
    else:
        print(f"Gaussian parameters not found for subject {name}")

# Save result workbook
output_book.save('2.Gaussian_optimal_parameters.xlsx')
output_book.close()