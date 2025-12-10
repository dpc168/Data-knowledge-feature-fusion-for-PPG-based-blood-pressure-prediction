import openpyxl

book1 = openpyxl.load_workbook("8D_KF.xlsx")
sheet1 = book1["Sheet1"]
sheet1_max_row = sheet1.max_row

# Create new Excel file and worksheet
book2 = openpyxl.Workbook()
sheet2 = book2.active
sheet2.title = "Guassian_boundary"  # Rename worksheet

# Add headers
sheet2.append(['Name', 'start_a1', 'start_a2', 'start_a3', 'start_b1', 'start_b2', 'start_b3',
               'start_c1', 'start_c2', 'start_c3', 'lower_a1','lower_a2', 'lower_a3', 'lower_b1',
               'lower_b2', 'lower_b3', 'lower_c1', 'lower_c2', 'lower_c3', 'upper_a1 ', 'upper_a2 ',
               'upper_a3 ', 'upper_b1 ', 'upper_b2 ', 'upper_b3 ', 'upper_c1 ', 'upper_c2 ','upper_c3'])

for row in range(2, sheet1_max_row + 1):
    Name = sheet1.cell(row, 1).value
    sheet2.cell(row, 1).value = Name
    T1 = sheet1.cell(row, 4).value
    h1 = sheet1.cell(row, 5).value
    T2 = sheet1.cell(row, 6).value
    h2 = sheet1.cell(row, 7).value
    Tsys = sheet1.cell(row, 8).value
    T3 = sheet1.cell(row, 9).value
    h3 = sheet1.cell(row, 10).value
    T = sheet1.cell(row, 11).value
    Tdia = sheet1.cell(row, 12).value

    # Gaussian iteration range
    # start
    start_a1 = 0.8 * h1
    start_a2 = 0.8 * h2
    start_a3 = 0.8 * h3
    start_b1 = T1
    start_b2 = T2
    start_b3 = T3
    start_c1 = Tsys / 6
    start_c2 = Tsys / 6
    start_c3 = Tdia / 6

    # Write
    sheet2.cell(row, 2).value = start_a1
    sheet2.cell(row, 3).value = start_a2
    sheet2.cell(row, 4).value = start_a3

    sheet2.cell(row, 5).value = start_b1
    sheet2.cell(row, 6).value = start_b2
    sheet2.cell(row, 7).value = start_b3

    sheet2.cell(row, 8).value = start_c1
    sheet2.cell(row, 9).value = start_c2
    sheet2.cell(row, 10).value = start_c3

    # lower
    lower_a1 = 0.8 * h1
    lower_a2 = 0.8 * h2
    lower_a3 = 0.8 * h3
    lower_b1 = T1 - 0.1 * T
    lower_b2 = T2 - 0.1 * T
    lower_b3 = T3 - 0.1 * T
    lower_c1 = 0.5 * Tsys / 6
    lower_c2 = 0.5 * Tsys / 6
    lower_c3 = 0.5 * Tdia / 6

    # Write
    sheet2.cell(row, 11).value = lower_a1
    sheet2.cell(row, 12).value = lower_a2
    sheet2.cell(row, 13).value = lower_a3

    sheet2.cell(row, 14).value = lower_b1
    sheet2.cell(row, 15).value = lower_b2
    sheet2.cell(row, 16).value = lower_b3

    sheet2.cell(row, 17).value = lower_c1
    sheet2.cell(row, 18).value = lower_c2
    sheet2.cell(row, 19).value = lower_c3

    # upper
    upper_a1 = h1
    upper_a2 = h2
    upper_a3 = h3
    upper_b1 = T1 + 0.1 * T
    upper_b2 = T2 + 0.1 * T
    upper_b3 = T3 + 0.1 * T
    upper_c1 = 1.5 * Tsys / 6
    upper_c2 = 1.5 * Tsys / 6
    upper_c3 = 1.5 * Tdia / 6
    # Write
    sheet2.cell(row, 20).value = upper_a1
    sheet2.cell(row, 21).value = upper_a2
    sheet2.cell(row, 22).value = upper_a3

    sheet2.cell(row, 23).value = upper_b1
    sheet2.cell(row, 24).value = upper_b2
    sheet2.cell(row, 25).value = upper_b3

    sheet2.cell(row, 26).value = upper_c1
    sheet2.cell(row, 27).value = upper_c2
    sheet2.cell(row, 28).value = upper_c3

book2.save("1_Guassian_boundary.xlsx")
book2.close()